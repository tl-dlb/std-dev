from datetime import datetime
from django.utils import timezone
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from tempfile import NamedTemporaryFile
from standard.models import Lot

tz = timezone.get_default_timezone()

def create_daily_report(start_date, end_date):
    start_date = datetime.strptime(start_date, "%d.%m.%Y").date()
    end_date = datetime.strptime(end_date, "%d.%m.%Y").date()
    lots = Lot.objects.filter(Q(status='COMPLETED') | Q(status='NOT_COMPLETED')).all()
    lots = lots.filter(bidding_begin__range=(start_date, end_date))

    wb = Workbook()
    ws = wb.active

    start = 1

    ws['A'+str(start)] = '№ п/п'  
    ws['B'+str(start)] = 'Время проведения, GMS Нур-Султан'
    ws['C'+str(start)] = 'Режим торгов'
    ws['D'+str(start)] = 'Номер сделки'
    ws['E'+str(start)] = 'Наименование'
    ws['F'+str(start)] = 'Код товара'
    
    ws['G'+str(start)] = 'Базис/условия поставки'
    ws['H'+str(start)] = 'Код ТН ВЭД'
    ws['I'+str(start)] = 'Цель торга'
    ws['J'+str(start)] = 'ПКО'

    ws['K'+str(start)] = 'Количество товара'
    ws['L'+str(start)] = 'Размер лота'
    ws['M'+str(start)] = 'Единица измерения лота'
    ws['N'+str(start)] = 'Наличие НДС'

    ws['O'+str(start)] = 'Первоначальная сумма, тенге'
    ws['P'+str(start)] = 'Цена за единицу, тенге'

    ws['Q'+str(start)] = 'Сумма сделки, тенге'
    ws['R'+str(start)] = 'Экономия, тенге'
    ws['S'+str(start)] = 'Экономия, %'
    ws['T'+str(start)] = 'Количество лотов'
    ws['U'+str(start)] = 'Количество участников аукциона'
    ws['V'+str(start)] = 'Квалификационные требования'

    ws['W'+str(start)]  = 'Наименование покупателя и (или) дилера'
    ws['X'+str(start)]  = 'БИН покупателя'
    ws['Y'+str(start)]  = 'Индивидуальный код клиринга покупателя'
    ws['Z'+str(start)]  = 'Наименование брокера покупателя'
    ws['AA'+str(start)] = 'БИН брокера покупателя'

    ws['AB'+str(start)] = 'Наименование продавца и (или) дилера'
    ws['AC'+str(start)] = 'БИН продавца'
    ws['AD'+str(start)] = 'Индивидуальный код клиринга продавца'
    ws['AE'+str(start)] = 'Наименование брокера'
    ws['AF'+str(start)] = 'БИН брокера'

    ws['AG'+str(start)] = 'Актуальный статус торга'

    i = start + 1
    for lot in lots:
        position = lot.positions.filter(status='ACTIVE').first()
        bid = lot.result.bid

        ws['A'+str(i)] = i - start
        ws['B'+str(i)] = lot.bidding_begin.astimezone(tz).strftime('%d.%m.%Y %H:%M')
        ws['C'+str(i)] = 'Стандартный аукцион на понижение' if lot.type == 'AUCTION_DOWN' else 'Стандартный аукцион на повышение'
        ws['D'+str(i)] = lot.number
        ws['E'+str(i)] = lot.name
        ws['F'+str(i)] = position.internal_code if position.internal_code else '--'

        ws['G'+str(i)] = position.delivery_terms
        ws['H'+str(i)] = position.external_code if position.external_code else '--'
        ws['I'+str(i)] = 'Покупка' if lot.type == 'AUCTION_DOWN' else 'Продажа'
        ws['J'+str(i)] = 'с ПКО' if lot.qualification else 'без ПКО'

        ws['K'+str(i)] = position.quantity
        ws['K'+str(i)].number_format = '#,##0.00'
        ws['L'+str(i)] = 1
        ws['M'+str(i)] = position.unit
        ws['N'+str(i)] = 'с НДС' if lot.vat else 'без НДС'

        ws['O'+str(i)] = lot.sum
        ws['O'+str(i)].number_format = '#,##0.00'
        ws['P'+str(i)] = position.price
        ws['P'+str(i)].number_format = '#,##0.00'

        ws['Q'+str(i)] = bid.sum if lot.status == 'COMPLETED' else '--'
        ws['Q'+str(i)].number_format = '#,##0.00'
        ws['R'+str(i)] = abs(lot.sum - bid.sum) if lot.status == 'COMPLETED' else '--'
        ws['R'+str(i)].number_format = '#,##0.00'
        ws['S'+str(i)] = abs(lot.sum - bid.sum) / lot.sum if lot.status == 'COMPLETED' else '--'
        ws['S'+str(i)].number_format = '0.00%'  
        ws['T'+str(i)] = 1
        ws['U'+str(i)] = lot.applications.filter(status='ADMITTED').count()
        ws['V'+str(i)] = '--'

        if lot.type == 'AUCTION_DOWN': 
            ws['W'+str(i)]  = lot.client.name 
            ws['X'+str(i)]  = lot.client.idn
            ws['Y'+str(i)]  = lot.client.clearing_code if lot.client.clearing_code else '--'
            ws['Z'+str(i)]  = lot.company.name
            ws['AA'+str(i)] = lot.company.idn

            if lot.status == 'COMPLETED':
                ws['AB'+str(i)] = bid.application.client.name 
                ws['AC'+str(i)] = bid.application.client.idn 
                ws['AD'+str(i)] = bid.application.client.clearing_code if bid.application.client.clearing_code else '--'
                ws['AE'+str(i)] = bid.application.company.name  
                ws['AF'+str(i)] = bid.application.company.idn    
        elif lot.type == 'AUCTION_UP': 
            if lot.status == 'COMPLETED':
                ws['W'+str(i)] = bid.application.client.name 
                ws['X'+str(i)] = bid.application.client.idn  
                ws['Y'+str(i)] = bid.application.client.clearing_code if bid.application.client.clearing_code else '--'
                ws['Z'+str(i)] = bid.application.company.name     
                ws['AA'+str(i)] = bid.application.company.idn       

            ws['AB'+str(i)]  = lot.client.name 
            ws['AC'+str(i)]  = lot.client.idn
            ws['AD'+str(i)]  = lot.client.clearing_code if lot.client.clearing_code else '--'
            ws['AE'+str(i)]  = lot.company.name
            ws['AF'+str(i)] = lot.company.idn

        ws['AG'+str(i)] = 'Завершен' if lot.status == 'COMPLETED' else 'Не состоялся'

        i = i + 1

    for row in ws['A1:AG'+str(i-1)]:
        for cell in row:
            cell.font = Font(name='Times New Roman')
            if cell.row >= start:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws['A'+str(start)+':AG'+str(start)]:
        for cell in row:
            cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 10
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 15
    ws.column_dimensions['Q'].width = 15
    ws.column_dimensions['R'].width = 15
    ws.column_dimensions['S'].width = 10
    ws.column_dimensions['T'].width = 10
    ws.column_dimensions['U'].width = 10
    ws.column_dimensions['V'].width = 10
    ws.column_dimensions['W'].width = 25
    ws.column_dimensions['X'].width = 15
    ws.column_dimensions['Y'].width = 10
    ws.column_dimensions['Z'].width = 25
    ws.column_dimensions['AA'].width = 15
    ws.column_dimensions['AB'].width = 25
    ws.column_dimensions['AC'].width = 15
    ws.column_dimensions['AD'].width = 10
    ws.column_dimensions['AE'].width = 25
    ws.column_dimensions['AF'].width = 15
    ws.column_dimensions['AG'].width = 15

    with NamedTemporaryFile(delete=False) as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    return stream
